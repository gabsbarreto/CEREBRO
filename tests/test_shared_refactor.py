from __future__ import annotations

import sys
import tempfile
import threading
import unittest
from pathlib import Path

from app import config
from app.models import JobSettings, public_model_presets
from app.services import jobs
from app.services import openai_rq
from app.shared.cli import parse_bool, parse_bool_arg
from app.shared.openai_responses import response_diagnostics, response_text
from app.shared.process_runner import run_event_process
from app.shared.rq_chat import build_chat_prompt, combined_prompt, strip_thinking


class SharedHelperTests(unittest.TestCase):
    def test_bool_parsing_preserves_script_semantics(self) -> None:
        self.assertTrue(parse_bool("true"))
        self.assertTrue(parse_bool("1"))
        self.assertFalse(parse_bool("false"))
        self.assertTrue(parse_bool_arg("yes"))
        self.assertFalse(parse_bool_arg("n"))

    def test_rq_chat_prompt_uses_chat_template_when_available(self) -> None:
        class Tokenizer:
            def apply_chat_template(self, messages, tokenize, add_generation_prompt, enable_thinking=False):
                self.messages = messages
                self.enable_thinking = enable_thinking
                return "templated"

        tokenizer = Tokenizer()
        self.assertEqual(build_chat_prompt(tokenizer, "SYS", "USER", True), "templated")
        self.assertEqual(
            tokenizer.messages,
            [{"role": "system", "content": "SYS"}, {"role": "user", "content": "USER"}],
        )
        self.assertTrue(tokenizer.enable_thinking)

    def test_rq_chat_prompt_fallback_and_thinking_strip(self) -> None:
        fallback = build_chat_prompt(object(), "SYS", "USER", False)
        self.assertEqual(fallback, combined_prompt("SYS", "USER"))
        self.assertEqual(strip_thinking("a <think>hidden</think> b"), "a  b")

    def test_openai_response_helpers_handle_sdk_shapes(self) -> None:
        class Content:
            type = "output_text"
            text = "from content"

        class Item:
            content = [Content()]
            type = "message"

        class Response:
            id = "resp_1"
            status = "completed"
            incomplete_details = None
            usage = {"total_tokens": 1}
            output = [Item()]

        response = Response()
        self.assertEqual(response_text(response), "from content")
        diagnostics = response_diagnostics(response)
        self.assertEqual(diagnostics["id"], "resp_1")
        self.assertEqual(diagnostics["output_types"], ["message"])

    def test_event_process_runner_parses_json_events(self) -> None:
        events = []
        result = run_event_process(
            cmd=[
                sys.executable,
                "-c",
                "import json; print('noise'); print(json.dumps({'event':'done','value':3}))",
            ],
            job_id=None,
            on_event=events.append,
            cancel_message="cancelled",
        )
        self.assertEqual(result.return_code, 0)
        self.assertIn("noise", result.details)
        self.assertEqual(events, [{"event": "done", "value": 3}])

    def test_openai_api_key_resolution_order(self) -> None:
        original_key_file = openai_rq.OPENAI_API_KEY_FILE
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                openai_rq.OPENAI_API_KEY_FILE = Path(tmpdir) / "api_key.txt"
                openai_rq.OPENAI_API_KEY_FILE.write_text("sk-file\n", encoding="utf-8")

                self.assertEqual(openai_rq.resolve_openai_api_key(" sk-form ", {}), "sk-form")
                self.assertEqual(
                    openai_rq.resolve_openai_api_key("", {"OPENAI_API_KEY": " sk-env "}),
                    "sk-env",
                )
                self.assertEqual(openai_rq.resolve_openai_api_key("", {}), "sk-file")

                openai_rq.OPENAI_API_KEY_FILE.unlink()
                self.assertEqual(openai_rq.resolve_openai_api_key("", {}), "")
        finally:
            openai_rq.OPENAI_API_KEY_FILE = original_key_file

    def test_gpt54_mini_high_preset_is_public_openai_model(self) -> None:
        settings = JobSettings.from_form({"rq_model_preset": "openai_gpt54_mini_high"})
        self.assertEqual(settings.rq_provider, "openai")
        self.assertEqual(settings.rq_screening_model, "gpt-5.4-mini")
        self.assertTrue(settings.rq_enable_thinking)
        self.assertEqual(settings.openai_reasoning_effort, "high")

        public_presets = {preset["id"]: preset for preset in public_model_presets()}
        self.assertIn("openai_gpt54_mini_high", public_presets)
        self.assertEqual(public_presets["openai_gpt54_mini_high"]["settings"]["model"], "gpt-5.4-mini")

    def test_openai_pdf_file_mode_is_openai_only(self) -> None:
        openai_settings = JobSettings.from_form(
            {"rq_model_preset": "openai_gpt5_mini_high", "openai_input_mode": "pdf_file"}
        )
        self.assertEqual(openai_settings.openai_input_mode, "pdf_file")

        local_settings = JobSettings.from_form(
            {"rq_model_preset": "qwen35_9b_8bit_reasoning", "openai_input_mode": "pdf_file"}
        )
        self.assertEqual(local_settings.openai_input_mode, "ocr_text")

    def test_openai_rq_passes_input_file_arguments(self) -> None:
        from app.services import openai_rq

        calls = []
        original_runner = openai_rq.run_event_process
        try:
            def fake_runner(**kwargs):
                calls.append(kwargs)
                return type("Result", (), {"return_code": 0, "details": ""})()

            openai_rq.run_event_process = fake_runner
            openai_rq.run_openai_rq(
                job_id="job-file",
                model="gpt-5.4-mini",
                system_prompt_file=Path("system.txt"),
                user_prompt_file=Path("user.txt"),
                output_file=Path("out.md"),
                max_tokens=100,
                enable_reasoning=True,
                reasoning_effort="high",
                input_file_id="file-abc",
            )
        finally:
            openai_rq.run_event_process = original_runner

        cmd = calls[0]["cmd"]
        self.assertIn("--input-file-id", cmd)
        self.assertEqual(cmd[cmd.index("--input-file-id") + 1], "file-abc")


class JobIdentityTests(unittest.TestCase):
    def test_settings_metadata_updates_preserve_persisted_shape(self) -> None:
        settings = JobSettings.from_form(
            {
                "rq_model_preset": "openai_gpt5_mini_high",
                "openai_api_key": "sk-secret",
                "rq_prompt_filename": "prompt_a.txt",
                "rq_system_prompt": "system",
            }
        )

        payload = jobs.settings_metadata_updates(settings)
        self.assertEqual(payload["rq_provider"], "openai")
        self.assertEqual(payload["rq_screening_model"], "gpt-5-mini")
        self.assertEqual(payload["rq_prompt_filename"], "prompt_a.txt")
        self.assertEqual(payload["openai_reasoning_effort"], "high")
        self.assertNotIn("openai_api_key", payload["settings"])
        self.assertTrue(payload["settings"]["openai_api_key_provided"])

        completion_payload = jobs.settings_metadata_updates(
            settings,
            rq_prompt_filename="loaded_prompt.txt",
            include_settings=False,
        )
        self.assertEqual(completion_payload["rq_prompt_filename"], "loaded_prompt.txt")
        self.assertNotIn("settings", completion_payload)

    def test_run_identity_and_ocr_copy_helpers(self) -> None:
        original_jobs_dir = config.JOBS_DIR
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                config.JOBS_DIR = Path(tmpdir) / "jobs"
                config.JOBS_DIR.mkdir(parents=True, exist_ok=True)
                source_id = "source"
                target_id = "target"
                source_root = jobs.create_job(
                    source_id,
                    "same-file.pdf",
                    JobSettings.from_form(
                        {"rq_prompt_filename": "prompt1.txt", "rq_model_preset": "openai_gpt5_mini_high"}
                    ),
                )
                (source_root / "outputs" / "rq_screening_output.md").write_text("decision", encoding="utf-8")
                (source_root / "outputs" / "merged_full_text.txt").write_text(
                    "merged OCR text long enough to reuse",
                    encoding="utf-8",
                )
                (source_root / "ocr_text" / "page_0001.md").write_text("page OCR", encoding="utf-8")
                jobs.update_metadata(
                    source_root,
                    rq_prompt_filename="prompt1.txt",
                    rq_screening_model="gpt-5-mini",
                    openai_input_mode="ocr_text",
                    number_of_pages=1,
                )
                jobs.update_status(source_id, status="complete", stage="complete", message="complete", progress=1.0)

                self.assertEqual(
                    jobs.find_screened_job_by_run_identity("same-file.pdf", "prompt1.txt", "gpt-5-mini")["job_id"],
                    source_id,
                )
                self.assertIsNone(jobs.find_screened_job_by_run_identity("same-file.pdf", "prompt2.txt", "gpt-5-mini"))
                self.assertIsNone(
                    jobs.find_screened_job_by_run_identity(
                        "same-file.pdf",
                        "prompt1.txt",
                        "gpt-5-mini",
                        "pdf_file",
                    )
                )

                target_root = jobs.create_job(target_id, "same-file.pdf", JobSettings())
                jobs.copy_reusable_ocr(source_id, target_root)
                self.assertTrue((target_root / "ocr_text" / "page_0001.md").exists())
                self.assertIn(
                    "merged OCR",
                    (target_root / "outputs" / "merged_full_text.txt").read_text(encoding="utf-8"),
                )
                self.assertTrue(jobs.read_metadata(target_root)["pending_rerun_screening_only"])
        finally:
            config.JOBS_DIR = original_jobs_dir

    def test_reusable_openai_file_id_helpers(self) -> None:
        original_jobs_dir = config.JOBS_DIR
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                config.JOBS_DIR = Path(tmpdir) / "jobs"
                config.JOBS_DIR.mkdir(parents=True, exist_ok=True)
                source_root = jobs.create_job("source-file", "Pilots/122710267.pdf", JobSettings())
                jobs.update_metadata(
                    source_root,
                    pdf_sha256="digest-1",
                    openai_file_id="file-abc",
                    openai_input_mode="pdf_file",
                )
                match = jobs.find_reusable_openai_file_job("digest-1", "Other/122710267.pdf")
                self.assertEqual(match["job_id"], "source-file")

                target_root = jobs.create_job("target-file", "122710267.pdf", JobSettings())
                jobs.copy_reusable_openai_file("source-file", target_root)
                metadata = jobs.read_metadata(target_root)
                self.assertEqual(metadata["openai_file_id"], "file-abc")
                self.assertEqual(metadata["openai_file_reused_from_job_id"], "source-file")
        finally:
            config.JOBS_DIR = original_jobs_dir

    def test_different_identity_rerun_child_job_reuses_source_ocr(self) -> None:
        original_jobs_dir = config.JOBS_DIR
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                config.JOBS_DIR = Path(tmpdir) / "jobs"
                config.JOBS_DIR.mkdir(parents=True, exist_ok=True)
                source_settings = JobSettings.from_form(
                    {"rq_prompt_filename": "prompt1.txt", "rq_model_preset": "openai_gpt5_mini_high"}
                )
                source_root = jobs.create_job("source", "same-file.pdf", source_settings)
                (source_root / "input" / "uploaded.pdf").write_bytes(b"%PDF source")
                (source_root / "outputs" / "merged_full_text.txt").write_text(
                    "merged OCR text long enough to reuse",
                    encoding="utf-8",
                )
                (source_root / "ocr_text" / "page_0001.md").write_text("page OCR", encoding="utf-8")
                jobs.update_metadata(
                    source_root,
                    rq_prompt_filename="prompt1.txt",
                    rq_screening_model="gpt-5-mini",
                    number_of_pages=1,
                )

                self.assertTrue(jobs.job_matches_run_identity(source_root, source_settings))
                child_settings = JobSettings.from_form(
                    {"rq_prompt_filename": "prompt2.txt", "rq_model_preset": "qwen35_9b_8bit_reasoning"}
                )
                self.assertFalse(jobs.job_matches_run_identity(source_root, child_settings))

                child_id, child_root = jobs.create_screening_rerun_child_job("source", child_settings)
                self.assertNotEqual(child_id, "source")
                self.assertEqual((child_root / "input" / "uploaded.pdf").read_bytes(), b"%PDF source")
                self.assertTrue((child_root / "ocr_text" / "page_0001.md").exists())
                self.assertIn(
                    "merged OCR",
                    (child_root / "outputs" / "merged_full_text.txt").read_text(encoding="utf-8"),
                )
                metadata = jobs.read_metadata(child_root)
                self.assertEqual(metadata["rerun_created_from_job_id"], "source")
                self.assertEqual(metadata["reused_ocr_from_job_id"], "source")
                self.assertEqual(metadata["rq_prompt_filename"], "prompt2.txt")
                self.assertEqual(metadata["rq_screening_model"], config.RQ_SCREENING_MODEL)
                self.assertEqual(metadata["rq_max_tokens"], 30000)
                self.assertTrue(metadata["rq_enable_thinking"])
                self.assertTrue(metadata["pending_rerun_screening_only"])
        finally:
            config.JOBS_DIR = original_jobs_dir


class PipelineConcurrencyTests(unittest.TestCase):
    def test_openai_pdf_file_mode_skips_ocr_and_stores_file_id(self) -> None:
        from app.services import rq_screening_pipeline as pipeline

        original_jobs_dir = config.JOBS_DIR
        original_summary_path = config.SUMMARY_XLSX_PATH
        original_page_count = pipeline.page_count
        original_render = pipeline.render_pdf_to_images
        original_ocr = pipeline.run_deepseek_ocr
        original_openai = pipeline.run_openai_rq
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                root = Path(tmpdir)
                config.JOBS_DIR = root / "jobs"
                config.SUMMARY_XLSX_PATH = root / "summary.xlsx"
                config.JOBS_DIR.mkdir(parents=True, exist_ok=True)

                pipeline.page_count = lambda _pdf_path: 2
                pipeline.render_pdf_to_images = lambda *args, **kwargs: self.fail("OCR rendering should be skipped")
                pipeline.run_deepseek_ocr = lambda *args, **kwargs: self.fail("OCR should be skipped")

                def fake_openai_runner(**kwargs) -> None:
                    self.assertEqual(kwargs["input_file_id"], "")
                    self.assertEqual(Path(kwargs["input_file_path"]).name, "uploaded.pdf")
                    callback = kwargs.get("on_event")
                    if callback is not None:
                        callback({"event": "openai_file_uploaded", "file_id": "file-abc"})
                        callback({"event": "rq_generation_finished"})
                    Path(kwargs["output_file"]).write_text("direct PDF result", encoding="utf-8")

                pipeline.run_openai_rq = fake_openai_runner
                settings = JobSettings.from_form(
                    {
                        "rq_model_preset": "openai_gpt5_mini_high",
                        "openai_input_mode": "pdf_file",
                        "rq_prompt_filename": "prompt_file.txt",
                        "rq_system_prompt": "Use the user-provided text.",
                    }
                )
                job_root = jobs.create_job("pdf-file-job", "paper.pdf", settings)
                (job_root / "input" / "uploaded.pdf").write_bytes(b"%PDF mock")

                pipeline.run_job("pdf-file-job", settings, defer_openai=False)

                status = jobs.read_status("pdf-file-job")
                metadata = jobs.read_metadata(job_root)
                self.assertEqual(status["status"], "complete")
                self.assertEqual(metadata["openai_file_id"], "file-abc")
                self.assertTrue(metadata["openai_file_complete"])
                self.assertFalse(metadata["ocr_complete"])
                self.assertFalse((job_root / "outputs" / "merged_full_text.txt").exists())
        finally:
            pipeline.page_count = original_page_count
            pipeline.render_pdf_to_images = original_render
            pipeline.run_deepseek_ocr = original_ocr
            pipeline.run_openai_rq = original_openai
            config.JOBS_DIR = original_jobs_dir
            config.SUMMARY_XLSX_PATH = original_summary_path

    def test_deferred_openai_inference_does_not_block_next_ocr_job(self) -> None:
        from app.services import rq_screening_pipeline as pipeline
        from app.services.openai_inference_queue import OpenAIInferenceQueue

        original_jobs_dir = config.JOBS_DIR
        original_summary_path = config.SUMMARY_XLSX_PATH
        original_queue = pipeline.openai_inference_queue
        original_page_count = pipeline.page_count
        original_render = pipeline.render_pdf_to_images
        original_discover = pipeline.discover_deepseek_model
        original_ocr = pipeline.run_deepseek_ocr

        events: list[str] = []
        event_lock = threading.Lock()
        first_openai_started = threading.Event()
        release_first_openai = threading.Event()

        def record(event: str) -> None:
            with event_lock:
                events.append(event)

        def fake_render(_pdf_path: Path, _rendered_dir: Path, ocr_dir: Path, *, dpi: int) -> list[Path]:
            image = ocr_dir / "page_0001.png"
            image.parent.mkdir(parents=True, exist_ok=True)
            image.write_bytes(b"image")
            return [image]

        def fake_ocr(**kwargs) -> None:
            job_id = kwargs["job_id"]
            output_dir = Path(kwargs["output_dir"])
            record(f"ocr_start:{job_id}")
            output_dir.mkdir(parents=True, exist_ok=True)
            (output_dir / "page_0001.md").write_text(
                f"OCR text for {job_id}. This is long enough to pass the merge validation.",
                encoding="utf-8",
            )
            callback = kwargs.get("on_event")
            if callback is not None:
                callback({"event": "ocr_finished"})
            record(f"ocr_finish:{job_id}")

        def fake_openai_runner(**kwargs) -> None:
            job_id = kwargs["job_id"]
            record(f"openai_start:{job_id}")
            if job_id == "job-1":
                first_openai_started.set()
                self.assertTrue(release_first_openai.wait(timeout=5), "Timed out waiting to release OpenAI mock")
            Path(kwargs["output_file"]).write_text(f"OpenAI result for {job_id}", encoding="utf-8")
            callback = kwargs.get("on_event")
            if callback is not None:
                callback({"event": "rq_generation_finished"})
            record(f"openai_finish:{job_id}")

        local_openai_queue = OpenAIInferenceQueue(
            max_workers=1,
            max_retries=0,
            runner=fake_openai_runner,
            sleep=lambda _delay: None,
        )

        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                root = Path(tmpdir)
                config.JOBS_DIR = root / "jobs"
                config.SUMMARY_XLSX_PATH = root / "summary.xlsx"
                config.JOBS_DIR.mkdir(parents=True, exist_ok=True)

                pipeline.openai_inference_queue = local_openai_queue
                pipeline.page_count = lambda _pdf_path: 1
                pipeline.render_pdf_to_images = fake_render
                pipeline.discover_deepseek_model = lambda: "mock-deepseek"
                pipeline.run_deepseek_ocr = fake_ocr

                settings = JobSettings.from_form(
                    {
                        "rq_model_preset": "openai_gpt5_mini_high",
                        "rq_prompt_filename": "prompt_async.txt",
                        "rq_system_prompt": "Use the user-provided text.",
                    }
                )
                for job_id in ["job-1", "job-2"]:
                    job_root = jobs.create_job(job_id, f"{job_id}.pdf", settings)
                    (job_root / "input" / "uploaded.pdf").write_bytes(b"%PDF mock")

                pipeline.run_job("job-1", settings, defer_openai=True)
                self.assertTrue(first_openai_started.wait(timeout=5), "OpenAI mock did not start")

                pipeline.run_job("job-2", settings, defer_openai=True)
                with event_lock:
                    snapshot = list(events)
                self.assertIn("ocr_start:job-2", snapshot)
                self.assertNotIn("openai_finish:job-1", snapshot)

                release_first_openai.set()
                local_openai_queue.join()

                self.assertEqual(jobs.read_status("job-1")["status"], "complete")
                self.assertEqual(jobs.read_status("job-2")["status"], "complete")
                self.assertEqual(
                    (jobs.job_dir("job-1") / "outputs" / "rq_screening_output.md").read_text(encoding="utf-8"),
                    "OpenAI result for job-1",
                )
                self.assertEqual(
                    (jobs.job_dir("job-2") / "outputs" / "rq_screening_output.md").read_text(encoding="utf-8"),
                    "OpenAI result for job-2",
                )
        finally:
            release_first_openai.set()
            local_openai_queue.shutdown()
            pipeline.openai_inference_queue = original_queue
            pipeline.page_count = original_page_count
            pipeline.render_pdf_to_images = original_render
            pipeline.discover_deepseek_model = original_discover
            pipeline.run_deepseek_ocr = original_ocr
            config.JOBS_DIR = original_jobs_dir
            config.SUMMARY_XLSX_PATH = original_summary_path

    def test_openai_queue_retries_transient_errors(self) -> None:
        from app.services.openai_inference_queue import OpenAIInferenceJob, OpenAIInferenceQueue

        original_jobs_dir = config.JOBS_DIR
        attempts = 0
        delays: list[float] = []
        completions: list[str] = []
        completion_started_at: list[float] = []

        def flaky_runner(**kwargs) -> None:
            nonlocal attempts
            attempts += 1
            if attempts == 1:
                raise RuntimeError("rate limit 429")
            Path(kwargs["output_file"]).write_text("ok", encoding="utf-8")

        def complete_stub(**kwargs) -> None:
            completions.append(kwargs["job_id"])
            completion_started_at.append(kwargs["started_at"])
            jobs.update_status(
                kwargs["job_id"],
                status="complete",
                stage="complete",
                message="done",
                progress=1.0,
                event={"event": "complete"},
            )

        retry_queue = OpenAIInferenceQueue(
            max_workers=1,
            max_retries=1,
            retry_base_seconds=0.01,
            runner=flaky_runner,
            completion_handler=complete_stub,
            sleep=delays.append,
            clock=lambda: 123.45,
        )
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                root = Path(tmpdir)
                config.JOBS_DIR = root / "jobs"
                config.JOBS_DIR.mkdir(parents=True, exist_ok=True)
                settings = JobSettings.from_form(
                    {
                        "rq_model_preset": "openai_gpt5_mini_high",
                        "rq_prompt_filename": "prompt_retry.txt",
                        "rq_system_prompt": "Use the user-provided text.",
                    }
                )
                job_root = jobs.create_job("retry-job", "retry.pdf", settings)
                system_prompt = job_root / "outputs" / "rq_system_prompt.txt"
                user_prompt = job_root / "outputs" / "rq_user_prompt.txt"
                output = job_root / "outputs" / "rq_screening_output.md"
                pdf_path = job_root / "input" / "uploaded.pdf"
                system_prompt.write_text("system", encoding="utf-8")
                user_prompt.write_text("user text", encoding="utf-8")
                pdf_path.write_bytes(b"%PDF mock")

                self.assertTrue(
                    retry_queue.enqueue(
                        OpenAIInferenceJob(
                            job_id="retry-job",
                            settings=settings,
                            started_at=0.0,
                            prompt_filename="prompt_retry.txt",
                            prompt_source_path="",
                            system_prompt_file=system_prompt,
                            user_prompt_file=user_prompt,
                            output_file=output,
                            pdf_path=pdf_path,
                        )
                    )
                )
                retry_queue.join()
                self.assertEqual(attempts, 2)
                self.assertEqual(delays, [0.01])
                self.assertEqual(completions, ["retry-job"])
                self.assertEqual(completion_started_at, [123.45])
                self.assertEqual(jobs.read_status("retry-job")["status"], "complete")
        finally:
            retry_queue.shutdown()
            config.JOBS_DIR = original_jobs_dir


class RerunEndpointTests(unittest.TestCase):
    def test_upload_display_name_keeps_folder_path_only_in_metadata(self) -> None:
        from app.main import source_folder_from_relative_path, upload_display_filename

        class Upload:
            filename = "PDFs included/122710267.pdf"

        self.assertEqual(upload_display_filename(Upload()), "122710267.pdf")
        self.assertEqual(source_folder_from_relative_path("PDFs included/122710267.pdf"), "PDFs included")
        self.assertEqual(source_folder_from_relative_path("122710267.pdf"), "")

    def test_rerun_with_different_identity_returns_new_job_record(self) -> None:
        from fastapi.testclient import TestClient
        from app.main import app, job_queue

        original_data_dir = config.DATA_DIR
        original_jobs_dir = config.JOBS_DIR
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                root = Path(tmpdir)
                config.DATA_DIR = root / "data"
                config.JOBS_DIR = config.DATA_DIR / "jobs"
                config.JOBS_DIR.mkdir(parents=True, exist_ok=True)
                job_queue.pause()

                source_settings = JobSettings.from_form(
                    {"rq_prompt_filename": "prompt1.txt", "rq_model_preset": "openai_gpt5_mini_high"}
                )
                source_root = jobs.create_job("source", "same-file.pdf", source_settings)
                (source_root / "input" / "uploaded.pdf").write_bytes(b"%PDF source")
                (source_root / "outputs" / "merged_full_text.txt").write_text(
                    "merged OCR text long enough to reuse",
                    encoding="utf-8",
                )
                (source_root / "ocr_text" / "page_0001.md").write_text("page OCR", encoding="utf-8")
                jobs.update_metadata(
                    source_root,
                    rq_prompt_filename="prompt1.txt",
                    rq_screening_model="gpt-5-mini",
                    number_of_pages=1,
                )
                jobs.update_status("source", status="complete", stage="complete", message="done", progress=1.0)

                response = TestClient(app).post(
                    "/api/jobs/source/rerun",
                    data={
                        "rq_model_preset": "qwen35_9b_8bit_reasoning",
                        "rq_prompt_filename": "prompt2.txt",
                        "rq_system_prompt": "Use the user-provided text.",
                    },
                )
                self.assertEqual(response.status_code, 200)
                payload = response.json()
                self.assertTrue(payload["created_new_job"])
                self.assertEqual(payload["source_job_id"], "source")
                self.assertNotEqual(payload["job"]["job_id"], "source")
                child_metadata = payload["job"]["metadata"]
                self.assertEqual(child_metadata["rerun_created_from_job_id"], "source")
                self.assertEqual(child_metadata["rq_prompt_filename"], "prompt2.txt")
                self.assertEqual(child_metadata["rq_screening_model"], config.RQ_SCREENING_MODEL)
                self.assertEqual(child_metadata["rq_max_tokens"], 30000)
                self.assertTrue(child_metadata["rq_enable_thinking"])

                job_queue.clean_queued()
                job_queue.resume()
        finally:
            config.DATA_DIR = original_data_dir
            config.JOBS_DIR = original_jobs_dir


class BackendSmokeTests(unittest.TestCase):
    def test_key_get_endpoints_respond(self) -> None:
        original_jobs_dir = config.JOBS_DIR
        original_prompts_dir = config.PROMPTS_DIR
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                root = Path(tmpdir)
                config.JOBS_DIR = root / "jobs"
                config.PROMPTS_DIR = root / "prompts"
                config.JOBS_DIR.mkdir(parents=True, exist_ok=True)
                config.PROMPTS_DIR.mkdir(parents=True, exist_ok=True)

                from fastapi.testclient import TestClient
                from app.main import app

                client = TestClient(app)
                expectations = {
                    "/": 307,
                    "/rq-screening": 200,
                    "/api/jobs": 200,
                    "/api/queue": 200,
                    "/api/rq-prompts": 200,
                }
                for path, expected_status in expectations.items():
                    with self.subTest(path=path):
                        response = client.get(path, follow_redirects=False)
                        self.assertEqual(response.status_code, expected_status)
        finally:
            config.JOBS_DIR = original_jobs_dir
            config.PROMPTS_DIR = original_prompts_dir


class ExcelSummaryTests(unittest.TestCase):
    def test_summary_append_migrates_existing_workbook_to_prompt_column(self) -> None:
        from openpyxl import Workbook, load_workbook

        from app.services.excel_summary import SUMMARY_HEADERS, append_summary_row

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "existing.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(
                [
                    "job_id",
                    "filename",
                    "when it was inferenced",
                    "how long it took",
                    "LLM model",
                    "LLM output",
                ]
            )
            worksheet.append(["old-job", "old.pdf", "old-date", 1.0, "old-model", "old-output"])
            workbook.save(workbook_path)

            append_summary_row(
                job_id="new-job",
                filename="new.pdf",
                inferenced_at="2026-05-20T00:00:00Z",
                duration_seconds=3.5,
                llm_model="gpt-5-mini",
                llm_output="new-output",
                prompt="prompt_a.txt",
                workbook_path=workbook_path,
            )

            workbook = load_workbook(workbook_path)
            worksheet = workbook.active
            self.assertEqual(
                [worksheet.cell(row=1, column=index).value for index in range(1, len(SUMMARY_HEADERS) + 1)],
                SUMMARY_HEADERS,
            )
            self.assertEqual(worksheet.cell(row=2, column=6).value, None)
            self.assertEqual(worksheet.cell(row=2, column=7).value, "old-output")
            self.assertEqual(worksheet.cell(row=3, column=6).value, "prompt_a.txt")
            self.assertEqual(worksheet.cell(row=3, column=7).value, "new-output")

    def test_summary_append_and_rebuild_share_headers_and_layout(self) -> None:
        from openpyxl import load_workbook

        from app.services.excel_summary import (
            SUMMARY_COLUMN_WIDTHS,
            SUMMARY_HEADERS,
            append_summary_row,
            rebuild_summary_from_jobs,
        )

        original_jobs_dir = config.JOBS_DIR
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                root = Path(tmpdir)
                append_path = root / "append.xlsx"
                append_summary_row(
                    job_id="job-1",
                    filename="Pilots/122710267.pdf",
                    inferenced_at="2026-05-20T00:00:00Z",
                    duration_seconds=3.5,
                    llm_model="gpt-5-mini",
                    llm_output="decision",
                    prompt="prompt_a.txt",
                    workbook_path=append_path,
                )
                workbook = load_workbook(append_path)
                worksheet = workbook.active
                self.assertEqual(
                    [worksheet.cell(row=1, column=index).value for index in range(1, len(SUMMARY_HEADERS) + 1)],
                    SUMMARY_HEADERS,
                )
                self.assertEqual(worksheet.cell(row=2, column=5).value, "gpt-5-mini")
                self.assertEqual(worksheet.cell(row=2, column=6).value, "prompt_a.txt")
                self.assertEqual(worksheet.cell(row=2, column=2).value, "122710267")
                for column, width in SUMMARY_COLUMN_WIDTHS.items():
                    self.assertEqual(worksheet.column_dimensions[column].width, width)

                config.JOBS_DIR = root / "jobs"
                config.JOBS_DIR.mkdir(parents=True, exist_ok=True)
                job_root = jobs.create_job("job-2", "Pilots/screened.pdf", JobSettings())
                (job_root / "outputs" / "rq_screening_output.md").write_text("screening output", encoding="utf-8")
                jobs.update_metadata(
                    job_root,
                    completed_at="2026-05-20T01:00:00Z",
                    duration_seconds=4.0,
                    rq_screening_model="qwen",
                    rq_prompt_filename="prompt_b.txt",
                )
                jobs.update_status("job-2", status="complete", stage="complete", message="done", progress=1.0)

                rebuild_path = root / "rebuild.xlsx"
                rebuild_summary_from_jobs(workbook_path=rebuild_path)
                workbook = load_workbook(rebuild_path)
                worksheet = workbook.active
                self.assertEqual(
                    [worksheet.cell(row=1, column=index).value for index in range(1, len(SUMMARY_HEADERS) + 1)],
                    SUMMARY_HEADERS,
                )
                self.assertEqual(worksheet.cell(row=2, column=2).value, "screened")
                self.assertEqual(worksheet.cell(row=2, column=5).value, "qwen")
                self.assertEqual(worksheet.cell(row=2, column=6).value, "prompt_b.txt")
                for column, width in SUMMARY_COLUMN_WIDTHS.items():
                    self.assertEqual(worksheet.column_dimensions[column].width, width)
        finally:
            config.JOBS_DIR = original_jobs_dir


if __name__ == "__main__":
    unittest.main()
