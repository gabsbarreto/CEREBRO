from __future__ import annotations

import threading
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app import config
from app.services import jobs

SUMMARY_HEADERS = [
    "job_id",
    "filename",
    "when it was inferenced",
    "how long it took",
    "LLM model",
    "prompt",
    "LLM output",
]
SUMMARY_COLUMN_WIDTHS = {
    "A": 34,
    "B": 42,
    "C": 24,
    "D": 18,
    "E": 44,
    "F": 34,
    "G": 100,
}
_SUMMARY_LOCK = threading.Lock()


def append_summary_row(
    *,
    job_id: str,
    filename: str,
    inferenced_at: str,
    duration_seconds: float,
    llm_model: str,
    llm_output: str,
    prompt: str = "",
    workbook_path: Path | None = None,
) -> Path:
    path = workbook_path or config.SUMMARY_XLSX_PATH
    path.parent.mkdir(parents=True, exist_ok=True)

    with _SUMMARY_LOCK:
        workbook = load_workbook(path) if path.exists() else Workbook()
        worksheet = workbook.active
        worksheet.title = "RQ Screening"
        ensure_headers(worksheet)
        worksheet.append(
            [
                job_id,
                summary_filename(filename),
                inferenced_at,
                duration_seconds,
                llm_model,
                prompt,
                llm_output,
            ]
        )
        apply_summary_layout(worksheet)
        workbook.save(path)
    return path


def rebuild_summary_from_jobs(workbook_path: Path | None = None) -> Path:
    path = workbook_path or config.SUMMARY_XLSX_PATH
    path.parent.mkdir(parents=True, exist_ok=True)

    rows: list[list[Any]] = []
    for record in reversed(jobs.list_jobs(limit=0)):
        status = record.get("status") or {}
        metadata = record.get("metadata") or {}
        if status.get("status") != "complete":
            continue
        root = jobs.job_dir(str(record["job_id"]))
        output_file = root / "outputs" / "rq_screening_output.md"
        if not output_file.exists():
            continue
        rows.append(
            [
                str(record["job_id"]),
                summary_filename(str(metadata.get("original_filename") or record.get("filename") or "")),
                str(metadata.get("completed_at") or ""),
                metadata.get("duration_seconds"),
                str(metadata.get("rq_screening_model") or ""),
                str(metadata.get("rq_prompt_filename") or ""),
                output_file.read_text(encoding="utf-8"),
            ]
        )

    with _SUMMARY_LOCK:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "RQ Screening"
        for index, header in enumerate(SUMMARY_HEADERS, start=1):
            worksheet.cell(row=1, column=index, value=header)
        for row in rows:
            worksheet.append(row)
        apply_summary_layout(worksheet)
        workbook.save(path)
    return path


def summary_filename(filename: str) -> str:
    name = str(filename or "").strip().replace("\\", "/").rstrip("/")
    if not name:
        return ""
    return Path(name).name.rsplit(".", maxsplit=1)[0]


def apply_summary_layout(worksheet: Worksheet) -> None:
    for column, width in SUMMARY_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column].width = width


def ensure_headers(worksheet: Worksheet) -> None:
    existing = [worksheet.cell(row=1, column=index).value for index in range(1, len(SUMMARY_HEADERS) + 1)]
    if existing == SUMMARY_HEADERS:
        return
    if existing[0:6] == [
        "job_id",
        "filename",
        "when it was inferenced",
        "how long it took",
        "LLM model",
        "LLM output",
    ]:
        worksheet.insert_cols(6)
        worksheet.cell(row=1, column=6, value="prompt")
        return
    if existing[0:5] == ["job_id", "filename", "when it was inferenced", "how long it took", "LLM output"]:
        worksheet.insert_cols(5, amount=2)
        worksheet.cell(row=1, column=5, value="LLM model")
        worksheet.cell(row=1, column=6, value="prompt")
        return
    if existing[0:4] == ["filename", "when it was inferenced", "how long it took", "LLM output"]:
        worksheet.insert_cols(1)
        worksheet.cell(row=1, column=1, value="job_id")
        worksheet.insert_cols(5, amount=2)
        worksheet.cell(row=1, column=5, value="LLM model")
        worksheet.cell(row=1, column=6, value="prompt")
        return
    if worksheet.max_row == 1 and all(value is None for value in existing):
        for index, header in enumerate(SUMMARY_HEADERS, start=1):
            worksheet.cell(row=1, column=index, value=header)
        return

    worksheet.insert_rows(1)
    for index, header in enumerate(SUMMARY_HEADERS, start=1):
        worksheet.cell(row=1, column=index, value=header)
